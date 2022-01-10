# Author : 柠檬班-亚萌
# Project : lemon87
# Time : 2022/1/5 20:13
# E-mail : 3343787213@qq.com
# Company : 湖南零檬信息技术有限公司
# Site : http://www.lemonban.com
# Forum : http://testingpai.com

'''
接口自动化的代码步骤：   -- 相当于执行阶段的步骤
1、准备好excel的测试用例，并使用代码自动读取出来   -- read_data()
2、发送接口的请求，获取实际结果   -- func()
3、实际结果  vs  预期结果
4、回写最终的结果到excel  -- write_data()
'''
import openpyxl
import requests

def read_data(filename,sheet):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet]
    max_row = sheet.max_row  # 获取最大行数
    cases = []  # 定义一个空列表，存储所有的测试用例
    for i in range(2,max_row+1,1):
        dict1 = dict(
        case_id = sheet.cell(row=i, column=1).value,  # 获取用例编号
        header_login = sheet.cell(row=i,column=5).value,  # 获取请求头
        url_login = sheet.cell(row=i,column=6).value,  # 获取接口地址
        body_login = sheet.cell(row=i,column=7).value,  # 获取请求体
        expect_login = sheet.cell(row=i, column=8).value)
        cases.append(dict1)  # 将一条一条的用例循环的追加到列表里面
    return cases


def write_data(filename,sheetname,row,column,final_result):
    work_book = openpyxl.load_workbook(filename,)
    sheet_name =work_book[sheetname]
    sheet_name.cell(row=row, column=column).value = final_result
    work_book.save(filename)

def func(url,body):
    header_login = {"X-Lemonban-Media-Type":"lemonban.v2",
              "Content-Type":"application/json"}
    res = requests.post(url=url,json=body,headers=header_login)
    return res.json()


def execute_function(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        # 从excel里面读取出来的数据，全都是str
        case_id = case['case_id']  # 取出用例编号
        header = case['header_login']  # 取出请求头
        url = case['url_login']  # 取出接口地址
        body = case['body_login']  # 取出请求体
        expect = case['expect_login'] # 取出预期结果
        body = eval(body)  # eval()函数 运行被字符串包裹的python表达式
        # 也可以理解为用eval函数去除掉字符串的外衣 -- 单引号
        expect = eval(expect)
        res = func(url=url,body=body)  # 传参进来，调用发送请求的函数发送请求
        expect_code = expect['code']  # 取出预期结果里面的code信息
        res_code = res['code']  # 取出实际结果里面的code信息
        print('预期结果为：{}'.format(expect_code))
        print('实际结果为：{}'.format(res_code))
        if expect_code == res_code:
            print('{}功能的第{}条用例执行通过！！'.format(sheetname,case_id))
            final_res = 'pass'
        else:
            print('{}功能的第{}条用例执行失败！！'.format(sheetname,case_id))
            final_res = 'NG'
        print('*' * 50)
        write_data(filename,sheetname,case_id+1,9,final_res) # 调用回写结果的函数，写入最终结果

execute_function('testcase_api_wuye.xlsx','login')
execute_function('testcase_api_wuye.xlsx','register')